"""
Excel export endpoint.

Generates a polished .xlsx data dictionary workbook on-demand
from the current metadata in PostgreSQL (reflects admin edits).

GET /api/export/{dict_id}/excel
"""

from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Dictionary

router = APIRouter()


def _generate_workbook(metadata: dict) -> BytesIO:
    """Generate an Excel workbook from metadata dict, return as BytesIO buffer."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Style constants
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
    BODY_FONT = Font(name="Arial", size=10)
    LINK_FONT = Font(name="Arial", size=10, color="2E75B6", underline="single")
    ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    def style_body_cell(cell, is_alt=False):
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL

    def truncate_sheet_name(name, max_len=31):
        return name[:max_len] if len(name) > max_len else name

    tables = metadata.get("tables", [])
    wb = Workbook()

    # ── Master Index ──
    ws = wb.active
    ws.title = "Master Index"

    headers = [
        "#", "Table Name", "Column Count", "Catalog", "Database",
        "Type", "Comment", "Created Time", "Last Access", "Created By",
        "Provider", "Owner", "Location"
    ]
    col_widths = [6, 40, 13, 15, 15, 12, 45, 25, 20, 15, 10, 30, 60]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))

    detail_keys = [
        None, None, None, "Catalog", "Database",
        "Type", "Comment", "Created Time", "Last Access", "Created By",
        "Provider", "Owner", "Location"
    ]

    for idx, table in enumerate(tables):
        row = idx + 2
        t_name = table["table_name"]
        col_count = len(table.get("columns", []))
        sheet_name = truncate_sheet_name(t_name)
        detail = table.get("detail", {})
        is_alt = idx % 2 == 1

        c = ws.cell(row=row, column=1, value=idx + 1)
        style_body_cell(c, is_alt)
        c.alignment = CENTER

        cell = ws.cell(row=row, column=2, value=t_name)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = LINK_FONT
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL

        c = ws.cell(row=row, column=3, value=col_count)
        style_body_cell(c, is_alt)
        c.alignment = CENTER

        for ci in range(3, len(detail_keys)):
            key = detail_keys[ci]
            val = detail.get(key, "") if key else ""
            c = ws.cell(row=row, column=ci + 1, value=val)
            style_body_cell(c, is_alt)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(tables) + 1}"

    # ── Per-table sheets ──
    for idx, table in enumerate(tables):
        t_name = table["table_name"]
        sheet_name = truncate_sheet_name(t_name)
        ws = wb.create_sheet(title=sheet_name)

        back_cell = ws.cell(row=1, column=1, value="\u2190 Back to Master Index")
        back_cell.hyperlink = "#'Master Index'!A1"
        back_cell.font = Font(name="Arial", size=10, color="2E75B6", underline="single")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

        col_headers = [
            "table_name", "column_name", "data_type", "is_nullable",
            "ordinal_position", "comment", "source", "description"
        ]
        col_ws = [30, 35, 20, 12, 12, 45, 25, 50]

        for i, (h, w) in enumerate(zip(col_headers, col_ws), 1):
            ws.cell(row=2, column=i, value=h)
            ws.column_dimensions[get_column_letter(i)].width = w
        style_header(ws, 2, len(col_headers))

        for cidx, col in enumerate(table.get("columns", [])):
            row = cidx + 3
            is_alt = cidx % 2 == 1

            values = [
                t_name,
                col.get("column_name", ""),
                col.get("data_type", ""),
                col.get("is_nullable", ""),
                col.get("ordinal_position", ""),
                col.get("comment", ""),
                col.get("source", ""),
                col.get("description", ""),
            ]

            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                style_body_cell(cell, is_alt)
                if ci in (4, 5):
                    cell.alignment = CENTER

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(col_headers))}{len(table.get('columns', [])) + 2}"

    # Write to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/{dict_id}/excel")
def export_excel(dict_id: int, db: Session = Depends(get_db)):
    """Generate and download an Excel data dictionary workbook."""
    d = db.query(Dictionary).filter(Dictionary.id == dict_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Dictionary not found")

    metadata = d.metadata_json
    schema_name = metadata.get("schema", d.name)
    buffer = _generate_workbook(metadata)

    filename = f"data_dictionary_{schema_name.replace('.', '_')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
