"""
Generate an Excel data dictionary workbook from metadata JSON.

Usage:
    from generators.generate_excel import build_workbook
    build_workbook("outputs/metadata.json", "outputs/data_dictionary.xlsx")
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# --- Style constants ---
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="2E75B6", underline="single")
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_body_cell(cell, is_alt=False):
    cell.font = BODY_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    if is_alt:
        cell.fill = ALT_FILL


def _truncate_sheet_name(name, max_len=31):
    """Excel sheet names max 31 chars."""
    return name[:max_len] if len(name) > max_len else name


def create_master_sheet(wb, tables_meta):
    ws = wb.active
    ws.title = "Master Index"

    headers = [
        "#", "Table Name", "Column Count", "Catalog", "Database",
        "Type", "Comment", "Created Time", "Last Access", "Created By",
        "Provider", "Owner", "Location"
    ]
    col_widths = [6, 40, 13, 15, 15, 12, 45, 25, 20, 15, 10, 30, 60]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws, 1, len(headers))

    # Map header names to the keys from DESCRIBE EXTENDED output
    detail_keys = [
        None, None, None, "Catalog", "Database",
        "Type", "Comment", "Created Time", "Last Access", "Created By",
        "Provider", "Owner", "Location"
    ]

    for idx, table in enumerate(tables_meta):
        row = idx + 2
        t_name = table["table_name"]
        col_count = len(table["columns"])
        sheet_name = _truncate_sheet_name(t_name)
        detail = table.get("detail", {})
        is_alt = idx % 2 == 1

        # Col 1: row number
        c = ws.cell(row=row, column=1, value=idx + 1)
        _style_body_cell(c, is_alt)
        c.alignment = CENTER

        # Col 2: table name with hyperlink
        cell = ws.cell(row=row, column=2, value=t_name)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = LINK_FONT
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL

        # Col 3: column count
        c = ws.cell(row=row, column=3, value=col_count)
        _style_body_cell(c, is_alt)
        c.alignment = CENTER

        # Cols 4+: detail fields
        for ci in range(3, len(detail_keys)):
            key = detail_keys[ci]
            val = detail.get(key, "") if key else ""
            c = ws.cell(row=row, column=ci + 1, value=val)
            _style_body_cell(c, is_alt)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(tables_meta) + 1}"


def create_table_sheet(wb, table_meta, table_index):
    t_name = table_meta["table_name"]
    sheet_name = _truncate_sheet_name(t_name)
    ws = wb.create_sheet(title=sheet_name)

    # Row 1: back link to Master Index
    back_cell = ws.cell(row=1, column=1, value="\u2190 Back to Master Index")
    back_cell.hyperlink = "#'Master Index'!A1"
    back_cell.font = Font(name="Arial", size=10, color="2E75B6", underline="single")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Row 2: column headers
    headers = [
        "table_name", "column_name", "data_type", "is_nullable",
        "ordinal_position", "comment", "source", "description"
    ]
    col_widths = [30, 35, 20, 12, 12, 45, 25, 50]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=2, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws, 2, len(headers))

    # Row 3+: column data
    for idx, col in enumerate(table_meta["columns"]):
        row = idx + 3
        is_alt = idx % 2 == 1

        values = [
            t_name,
            col.get("column_name", ""),
            col.get("data_type", ""),
            col.get("is_nullable", ""),
            col.get("ordinal_position", ""),
            col.get("comment", ""),       # from DESCRIBE or information_schema
            "",                            # source — manual entry
            "",                            # description — manual / AI-assisted
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            _style_body_cell(cell, is_alt)
            if ci in (4, 5):
                cell.alignment = CENTER

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(table_meta['columns']) + 2}"


def build_workbook(metadata_path, output_path):
    with open(metadata_path) as f:
        metadata = json.load(f)

    tables = metadata["tables"]
    wb = Workbook()

    create_master_sheet(wb, tables)

    for idx, table in enumerate(tables):
        create_table_sheet(wb, table, idx)
        
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Data dictionary saved: {output_path}")
    print(f"  Sheets: 1 master + {len(tables)} table sheets")


if __name__ == "__main__":
    import sys
    meta = sys.argv[1] if len(sys.argv) > 1 else "outputs/metadata.json"
    out = sys.argv[2] if len(sys.argv) > 2 else "outputs/data_dictionary.xlsx"
    build_workbook(meta, out)
